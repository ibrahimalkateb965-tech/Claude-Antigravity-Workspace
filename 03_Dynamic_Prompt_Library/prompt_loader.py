import os
import openpyxl

class DynamicPromptLibrary:
    """
    Loads and searches prompt templates from G:\\أرشيف_ونسخ_احتياطية\\ملفات_مكررة_وعامة\\AI\\مكتبة الأوامر.xlsx
    dynamically using simple keyword search.
    """
    def __init__(self, excel_path=None):
        if excel_path is None:
            self.excel_path = r"G:\أرشيف_ونسخ_احتياطية\ملفات_مكررة_وعامة\AI\مكتبة الأوامر.xlsx"
        else:
            self.excel_path = excel_path
            
    def search_prompt(self, query):
        """
        Search for the most relevant prompt template based on query keywords.
        """
        if not os.path.exists(self.excel_path):
            return {"error": f"Prompt library file not found at: {self.excel_path}"}
            
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
            
        query_words = [w.lower() for w in query.split() if len(w) > 2]
        if not query_words:
            # Fallback if query is too short
            query_words = [query.lower()]
            
        matches = []
        
        for name in wb.sheetnames:
            ws = wb[name]
            for r in range(2, ws.max_row + 1):
                p_name = ws.cell(row=r, column=2).value
                p_text = ws.cell(row=r, column=3).value
                p_note = ws.cell(row=r, column=4).value if ws.max_column >= 4 else ""
                
                if p_name and p_text:
                    p_name_str = str(p_name).lower()
                    p_text_str = str(p_text).lower()
                    
                    # Score match based on keyword occurrences
                    score = 0
                    for word in query_words:
                        if word in p_name_str:
                            score += 5  # Higher weight for match in name/title
                        if word in p_text_str:
                            score += 1
                            
                    if score > 0:
                        matches.append({
                            "sheet": name,
                            "name": p_name,
                            "prompt": p_text,
                            "notes": p_note,
                            "score": score
                        })
                        
        wb.close()
        
        if not matches:
            return {"message": "No specific prompt template matched your query."}
            
        # Sort matches by score descending
        matches.sort(key=lambda x: x["score"], reverse=True)
        return {"best_match": matches[0], "all_matches_count": len(matches)}

if __name__ == "__main__":
    loader = DynamicPromptLibrary()
    # Test search
    print(loader.search_prompt("growth business"))
